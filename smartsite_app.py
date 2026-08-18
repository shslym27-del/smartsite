import streamlit as st
import pandas as pd
import json
import os
from datetime import date, datetime, timedelta
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SmartSite - 관급 공무 종합 생성기", layout="wide")

st.title("🏗️ SmartSite - 건설 공무 통합 서식 자동 생성 시스템")
st.caption(
    "관급공사 착공계 원본 서식(공문·표지·목차 포함)에 실제 값을 채워 넣고, "
    "법정 기준과 대조 검증한 뒤 제출용 엑셀 파일로 내려받는 도구입니다."
)
st.info(
    "🔒 이 서버는 입력하신 내용을 저장하지 않습니다. 현장 템플릿은 '내보내기'로 내 PC에 파일로 "
    "저장했다가, 다음 방문 때 '가져오기'로 업로드해서 이어서 쓰세요.",
    icon="🔒",
)

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "smartsite_template_master.xlsx")
GISUNG_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gisung_template_master.xlsx")

# ============================================================
# 0. 한글 금액 표기 (NUMBERSTRING 대체 — LibreOffice 미지원 함수 우회)
# ============================================================

def num_to_kr(n):
    """정수를 한자어 한글 표기로 변환 (NUMBERSTRING(n,1)과 동일 형식)."""
    n = int(n)
    if n == 0:
        return "영"
    digits = "일이삼사오육칠팔구"
    units_small = ["", "십", "백", "천"]
    units_big = ["", "만", "억", "조", "경"]
    chunks = []
    temp = n
    while temp > 0:
        chunks.append(temp % 10000)
        temp //= 10000
    result = ""
    for i in range(len(chunks) - 1, -1, -1):
        chunk = chunks[i]
        if chunk == 0:
            continue
        chunk_str = ""
        t = chunk
        for j in range(3, -1, -1):
            d = (t // (10 ** j)) % 10
            if d == 0:
                continue
            if d == 1 and j > 0:
                chunk_str += units_small[j]
            else:
                chunk_str += digits[d - 1] + units_small[j]
        result += chunk_str + units_big[i]
    return result


def amount_in_words(n):
    return f"일금 {num_to_kr(n)}원정 (₩{int(n):,}) "


# ============================================================
# 1. 법정 기준 함수 / 데이터
# ============================================================

def get_direct_construction_ratio(labor_amount):
    """
    건설산업기본법 제28조의2, 시행령 제30조의2 (2019.3.26 개정 기준)
    ※ 총 노무비 기준으로 판정 (도급금액 기준이 아님)
    ※ 도급금액 100억원 이하 공사에만 적용, 70억원 이상 구간은 별도 확인 필요
    """
    if labor_amount < 300_000_000:
        return 0.50, "노무비 기준 3억원 미만"
    elif labor_amount < 1_000_000_000:
        return 0.30, "노무비 기준 3억원 이상 ~ 10억원 미만"
    elif labor_amount < 3_000_000_000:
        return 0.20, "노무비 기준 10억원 이상 ~ 30억원 미만"
    elif labor_amount < 7_000_000_000:
        return 0.10, "노무비 기준 30억원 이상 ~ 70억원 미만"
    else:
        return None, "노무비 기준 70억원 이상 — 별표1 상한 구간 밖, 고시 원문 재확인 필요"


SAFETY_RATE_TABLE = {
    "일반건축공사": {"below5": 0.0311, "mid_rate": 0.0228, "mid_base": 4_325_000, "above50": 0.0237, "verified": True},
    "특수건설공사(전기·통신)": {"below5": 0.0207, "mid_rate": 0.0159, "mid_base": 2_450_000, "above50": 0.0164, "verified": True},
    "토목공사": {"below5": None, "mid_rate": None, "mid_base": None, "above50": None, "verified": False},
    "중건설공사": {"below5": None, "mid_rate": None, "mid_base": None, "above50": None, "verified": False},
}


def calc_safety_fee(target_amount, work_type):
    rates = SAFETY_RATE_TABLE[work_type]
    if rates["below5"] is None:
        return None, "이 공종은 요율이 미입력 상태입니다. 고시 원문 별표1을 확인해 아래에서 직접 입력해 주세요."
    if target_amount < 500_000_000:
        return target_amount * rates["below5"], None
    elif target_amount < 5_000_000_000:
        return target_amount * rates["mid_rate"] + rates["mid_base"], None
    else:
        return target_amount * rates["above50"], None


def fmt(n):
    if n is None:
        return "-"
    return f"{int(round(n)):,} 원"


# ============================================================
# 2. 사이드바 - 마스터 정보 (원본 '공사정보_출력X' 시트와 1:1 매핑)
# ============================================================

st.sidebar.header("📋 프로젝트 기본 정보")
st.sidebar.caption("여기 입력한 값은 착공계 전체 서식(공문·표지·목차 포함)에 자동으로 연결됩니다.")

project_name = st.sidebar.text_input("공사명", "")
work_category = st.sidebar.text_input("공사 공종", "종합공사")
site_address = st.sidebar.text_input("현장소재지", "")
report_date = st.sidebar.date_input("착공신고일자", value=date.today())

st.sidebar.markdown("**원청사**")
contractor_name = st.sidebar.text_input("상호", "")
contractor_tel = st.sidebar.text_input("전화번호", "")
contractor_addr = st.sidebar.text_input("소재지(주소)", "")
ceo_name = st.sidebar.text_input("대표자명", "")

st.sidebar.markdown("**발주처**")
client_name = st.sidebar.text_input("발주처명", "")
client_tel = st.sidebar.text_input("발주처 전화번호", "")

contract_number = st.sidebar.text_input("계약번호", "")
total_contract = st.sidebar.number_input("도급 계약 금액 (원)", value=0, step=100_000)
contract_date = st.sidebar.date_input("계약년월일", value=date.today())
start_date_master = st.sidebar.date_input("착공년월일", value=date.today())
end_date_master = st.sidebar.date_input("준공(예정)년월일", value=date.today() + timedelta(days=45))

st.sidebar.markdown("**현장대리인**")
site_agent = st.sidebar.text_input("성명", "")
site_agent_tel = st.sidebar.text_input("연락처", "")
site_agent_addr = st.sidebar.text_input("주소", "")
site_agent_license_no = st.sidebar.text_input("자격번호", "")
site_agent_grade = st.sidebar.selectbox("면허종류(등급)", ["특급", "고급", "중급", "초급"], index=2)

if total_contract == 0:
    st.sidebar.caption("👈 도급 계약 금액을 입력하면 각 탭의 자동 계산이 시작됩니다.")

contract_date_dt = datetime.combine(contract_date, datetime.min.time())

tabs = st.tabs([
    "🦺 산업안전보건관리비",
    "🌿 환경보전비",
    "🏗️ 직접시공계획서",
    "📝 착공신고서",
    "💰 선금",
    "📄 기성금청구/기성내역서",
    "🤝 하도급관리계획서·통보",
    "📦 착공계 패키지 다운로드",
])

# ============================================================
# TAB 1. 산업안전보건관리비
# ============================================================
with tabs[0]:
    st.subheader("산업안전보건관리비 자동 계상 및 검증")
    st.markdown(
        "산안비 = **대상액 × 요율** (+ 5억~50억 구간은 기초액 가산). "
        "대상액이 불분명하면 (총공사금액+부가세) × 70% 를 관행상 대상액으로 사용합니다."
    )

    work_type = st.selectbox("공사 종류", list(SAFETY_RATE_TABLE.keys()), key="safety_work_type")
    rates = SAFETY_RATE_TABLE[work_type]

    if not rates["verified"]:
        st.warning(
            f"⚠️ '{work_type}' 요율은 아직 검증되지 않았습니다. 고용노동부고시 「건설업 산업안전보건관리비 "
            "계상 및 사용기준」 별표1 원문을 확인한 뒤 아래 값을 직접 입력해 주세요."
        )
        c1, c2, c3 = st.columns(3)
        rates["below5"] = c1.number_input("5억원 미만 요율(%)", value=0.0, step=0.01, key="r1") / 100
        rates["mid_rate"] = c2.number_input("5~50억원 요율(%)", value=0.0, step=0.01, key="r2") / 100
        rates["mid_base"] = c3.number_input("5~50억원 기초액(원)", value=0, step=10_000, key="r3")
        rates["above50"] = st.number_input("50억원 이상 요율(%)", value=0.0, step=0.01, key="r4") / 100
    else:
        st.caption(
            f"적용 요율 — 5억 미만 {rates['below5']*100:.2f}% / "
            f"5~50억 {rates['mid_rate']*100:.2f}%+{rates['mid_base']:,}원 / "
            f"50억 이상 {rates['above50']*100:.2f}%  "
            "(2차 출처 기반, 제출 전 고시 원문 대조 권장)"
        )

    target_basis = st.radio("대상액 산정 방식", ["직접 입력", "총공사금액 × 70% 관행 산식"], horizontal=True)
    if target_basis == "총공사금액 × 70% 관행 산식":
        target_amount = int(total_contract * 1.1 * 0.7)
        st.metric("추정 대상액", fmt(target_amount))
    else:
        target_amount = st.number_input("대상액 (직접재료비+간접재료비+직접노무비, 원)", value=int(total_contract * 0.6), step=100_000)

    legal_fee, err = calc_safety_fee(target_amount, work_type)
    user_safety_fee = st.number_input("원가계산서상 확정 산안비 계상액 (원)", value=0, step=1_000, key="s_target")

    st.divider()
    c1, c2 = st.columns(2)
    c1.metric("법정 최소 계상액(계산값)", fmt(legal_fee) if legal_fee else "요율 미입력")
    c2.metric("확정 계상액", fmt(user_safety_fee))

    if err:
        st.info(err)
    elif legal_fee is not None:
        diff = user_safety_fee - legal_fee
        if diff < 0:
            st.error(f"❌ 확정 계상액이 법정 최소 계상액보다 {fmt(-diff)} 부족합니다.")
        else:
            st.success(f"✅ 법정 최소 계상액을 충족합니다. (여유분 {fmt(diff)})")

    st.divider()
    st.markdown("**산업안전보건관리비 사용계획 (항목별)** — 원본 「총괄 산업안전보건관리비 사용계획서」 표와 동일 항목")
    safety_items = [
        "1. 안전담당자 인건비(수당)및 안전보조원 인건비",
        "2. 안전시설비등",
        "3. 개인보호구 및 안전장비구입비등",
        "4. 안전진단비등",
        "5. 안전보건교육비 및 행사비등",
        "6. 근로자건강진단비등",
    ]
    safety_usage = []
    cols = st.columns(2)
    for i, item in enumerate(safety_items):
        with cols[i % 2]:
            v = st.number_input(item, value=0, step=10_000, key=f"safety_item_{i}")
            safety_usage.append(v)
    safety_usage_total = sum(safety_usage)
    st.metric("사용계획 합계", fmt(safety_usage_total))
    if user_safety_fee and safety_usage_total != user_safety_fee:
        st.warning(f"⚠️ 사용계획 합계가 확정 계상액과 {fmt(abs(safety_usage_total - user_safety_fee))} 차이납니다.")

# ============================================================
# TAB 2. 환경보전비
# ============================================================
with tabs[1]:
    st.subheader("환경보전비 입력 및 상식적 범위 점검")
    st.info(
        "환경보전비는 산안비처럼 고정 % 요율표가 있는 항목이 아니라, 공종별 물량(가설방음벽, 세륜시설, "
        "폐기물 처리 등)을 산출내역서에서 직접 집계하는 항목입니다. 이 탭은 법정 계산기가 아니라 "
        "**입력값이 통상적인 범위에서 벗어나는지 1차 점검**만 수행합니다."
    )
    e_target = st.number_input("원가계산서상 확정 환경보전비 계상액 (원)", value=0, step=1_000, key="e_target")
    ratio = e_target / total_contract * 100 if total_contract else 0
    st.metric("도급금액 대비 환경보전비 비율", f"{ratio:.3f}%")

    if ratio < 0.05:
        st.warning("⚠️ 도급금액 대비 비율이 매우 낮습니다(관행상 0.1~1% 내외). 산출내역서 누락 여부를 확인해 보세요.")
    elif ratio > 3:
        st.warning("⚠️ 도급금액 대비 비율이 이례적으로 높습니다. 과다 계상 여부를 확인해 보세요.")
    else:
        st.success("✅ 통상적인 범위 내에 있습니다. (참고용 휴리스틱이며 법적 기준이 아닙니다)")

    st.divider()
    st.markdown("**환경보전비 사용계획 (항목별)** — 원본 「환경보전비 사용계획서」 표와 동일 항목")
    env_items = [
        "1. 환경보전시설 설치 및 운영비용",
        "2. 환경계측비용",
        "3. 환경자료 및 홍보물 구입비용",
        "4. 환경관련 인건비용",
        "5. 환경교육 비용",
    ]
    env_usage = []
    cols = st.columns(2)
    for i, item in enumerate(env_items):
        with cols[i % 2]:
            v = st.number_input(item, value=0, step=10_000, key=f"env_item_{i}")
            env_usage.append(v)
    env_usage_total = sum(env_usage)
    st.metric("사용계획 합계", fmt(env_usage_total))
    if e_target and env_usage_total != e_target:
        st.warning(f"⚠️ 사용계획 합계가 확정 계상액과 {fmt(abs(env_usage_total - e_target))} 차이납니다.")

# ============================================================
# TAB 3. 직접시공계획서
# ============================================================
with tabs[2]:
    st.subheader("🏗️ 직접시공계획서 (노무비 기준, 법정 검증 포함)")
    st.caption(
        "원본 「10-1.직접시공내역서」 표와 동일하게 공종 6개 행으로 고정되어 있습니다. "
        "행 수를 늘리고 싶다면 알려주세요 — 원본 서식 구조를 함께 조정해야 합니다."
    )

    st.markdown("**현장별 템플릿 (파일로 저장·불러오기)**")
    c1, c2 = st.columns(2)
    with c1:
        uploaded_tpl = st.file_uploader("📂 템플릿 파일 가져오기(.json)", type="json", key="tpl_uploader")
        if uploaded_tpl is not None:
            try:
                st.session_state["direct_df"] = pd.DataFrame(json.load(uploaded_tpl))
                st.success("템플릿을 불러왔습니다.")
            except Exception:
                st.error("파일을 읽을 수 없습니다. SmartSite에서 내보낸 .json 파일인지 확인해 주세요.")

    default_rows = [
        {"공종명": "", "재료비": 0, "노무비": 0, "경비": 0, "직접시공 여부": True} for _ in range(6)
    ]
    if "direct_df" not in st.session_state:
        st.session_state["direct_df"] = pd.DataFrame(default_rows)

    edited_df = st.data_editor(st.session_state["direct_df"], num_rows="fixed", key="direct_editor")
    st.session_state["direct_df"] = edited_df

    with c2:
        tpl_json = json.dumps(edited_df.to_dict(orient="records"), ensure_ascii=False, indent=2)
        tpl_filename = f"{(project_name or '현장')}_직접시공템플릿.json"
        st.download_button(
            "💾 현재 표를 템플릿 파일로 내보내기",
            data=tpl_json.encode("utf-8"),
            file_name=tpl_filename,
            mime="application/json",
        )

    st.divider()

    total_labor = edited_df["노무비"].sum()
    direct_labor = edited_df.loc[edited_df["직접시공 여부"] == True, "노무비"].sum()

    min_ratio, ratio_desc = get_direct_construction_ratio(total_labor)

    c1, c2, c3 = st.columns(3)
    c1.metric("총 노무비 합계", fmt(total_labor))
    c2.metric("직접시공 체크된 노무비 합계", fmt(direct_labor))
    if min_ratio is not None:
        min_direct_labor = total_labor * min_ratio
        c3.metric("법정 최소 직접시공 노무비", fmt(min_direct_labor))
        st.caption(f"⚖️ 판정 구간: {ratio_desc} → 법정 최소 비율 {min_ratio*100:.0f}%")
        if direct_labor >= min_direct_labor:
            st.success(f"✅ 법정 기준을 충족합니다. (현재 비율 {direct_labor/total_labor*100 if total_labor else 0:.1f}%)")
        else:
            shortfall = min_direct_labor - direct_labor
            st.error(f"❌ 법정 기준에 {fmt(shortfall)} 부족합니다. (현재 비율 {direct_labor/total_labor*100 if total_labor else 0:.1f}%)")
    else:
        st.warning(f"⚖️ {ratio_desc}")

# ============================================================
# TAB 4. 착공신고서
# ============================================================
with tabs[3]:
    st.subheader("📝 착공신고서 체크리스트 및 기한 관리")
    st.caption("공사명·계약정보·현장대리인 정보는 왼쪽 사이드바 입력값을 그대로 사용합니다.")

    has_supervisor = st.checkbox("감리자 지정 공사임")

    kiscon_deadline = contract_date + timedelta(days=30)
    st.metric("KISCON 건설공사대장 통보 기한", kiscon_deadline.strftime("%Y-%m-%d"),
              f"계약체결일 +30일 (D-{(kiscon_deadline - date.today()).days})")

    st.markdown("**제출 서류 체크리스트**")
    checklist_items = [
        "발주처 착공신고서",
        "세움터(건축행정시스템) 착공신고",
        "현장대리인 지정(배치)신고서",
        "산업안전보건관리비 사용계획서",
        "환경보전비 사용계획서",
        "직접시공계획서",
        "안전관리계획서(해당 공사 시)",
        "품질관리계획서/품질시험계획서",
        "공사예정공정표",
    ]
    checked = {}
    for item in checklist_items:
        checked[item] = st.checkbox(item, key=f"chk_{item}")
    done = sum(checked.values())
    st.progress(done / len(checklist_items))
    st.caption(f"{done} / {len(checklist_items)} 항목 준비 완료")

# ============================================================
# TAB 5. 선금
# ============================================================
with tabs[4]:
    st.subheader("💰 선금 지급 및 정산 관리")
    advance_rate = st.slider("선금 지급률(%)", 0, 70, 30)
    advance_amount = int(total_contract * advance_rate / 100)
    st.metric("선금 지급 예정액", fmt(advance_amount))

    usage_deadline = contract_date + timedelta(days=14)
    st.caption(f"공사계약일반조건 기준 선금 사용기한(참고): 배분 통지일로부터 14일 이내 사용 개시 → {usage_deadline.strftime('%Y-%m-%d')} 경 확인 필요")

    st.write("▼ **선금 사용내역 입력**")
    if "advance_df" not in st.session_state:
        st.session_state["advance_df"] = pd.DataFrame([
            {"항목": "노무비 지급", "금액": 0},
            {"항목": "자재비 지급", "금액": 0},
            {"항목": "장비비 지급", "금액": 0},
        ])
    advance_df = st.data_editor(st.session_state["advance_df"], num_rows="dynamic", key="advance_editor")
    st.session_state["advance_df"] = advance_df

    used = advance_df["금액"].sum()
    remaining = advance_amount - used
    c1, c2 = st.columns(2)
    c1.metric("사용 내역 합계", fmt(used))
    c2.metric("선금 잔액", fmt(remaining), delta_color="inverse" if remaining < 0 else "normal")
    if remaining < 0:
        st.error("❌ 사용 내역 합계가 선금 지급액을 초과했습니다.")

    st.divider()
    labor_received = st.number_input("노무비 수령액 누계(전용계좌, 원)", value=0, step=100_000, key="labor_received")
    advance_receipt_date = st.date_input("선금 수령일", value=date.today(), key="advance_receipt_date")

# ============================================================
# TAB 6. 기성금청구 / 기성내역서
# ============================================================
with tabs[5]:
    st.subheader("📄 기성금청구 및 기성내역서")
    current_progress_pct = st.slider("현재 누계 공정율(%)", 0, 100, 30)
    prev_gisung_amount = st.number_input("전회기성금액(원)", value=0, step=100_000, key="prev_gisung_amount")
    progress_amount = st.number_input(
        "금회기성금액(원)", value=int(total_contract * current_progress_pct / 100) - prev_gisung_amount,
        step=100_000, key="progress_amount_input",
    )
    cumulative_gisung = prev_gisung_amount + progress_amount

    prev_advance = st.number_input("기 지급된 선금 총액(원)", value=0, step=100_000, key="prev_advance_input")
    advance_deduction = int(progress_amount * (prev_advance / total_contract)) if total_contract else 0
    gisung_deduction = st.number_input("기성금 공제액(하자보증금 등, 원)", value=0, step=100_000, key="gisung_deduction")
    labor_direct_pay = st.number_input("이번 회차 노무비(전용계좌 직접 지급분, 원)", value=0, step=100_000)
    payable = progress_amount - advance_deduction - gisung_deduction

    c1, c2, c3 = st.columns(3)
    c1.metric("금회기성금액", fmt(progress_amount))
    c2.metric("누계기성금액", fmt(cumulative_gisung))
    c3.metric("실 지급 예정액", fmt(payable))

    st.caption(
        "※ 선금 정산은 계약조건(균등정산/기성고 비례정산 등)에 따라 방식이 다를 수 있으므로 "
        "계약서상 선금정산 특약을 반드시 확인하세요. 노무비는 전용계좌를 통한 별도 직접 지급 대상입니다."
    )
    if labor_direct_pay > progress_amount:
        st.error("❌ 이번 회차 노무비가 기성금액을 초과합니다. 입력값을 확인하세요.")

    st.divider()
    st.markdown("**입금 계좌 정보**")
    c1, c2 = st.columns(2)
    bank_name = c1.text_input("입금은행", "")
    bank_account = c2.text_input("계좌번호", "")

    st.divider()
    st.markdown("**공동도급 여부**")
    is_jv = st.checkbox("이 공사는 공동도급(공동수급체) 공사입니다", value=False)
    jv_name = ""
    jv_share = 0
    if is_jv:
        c1, c2 = st.columns(2)
        jv_name = c1.text_input("공동수급사 상호", "")
        jv_share = c2.number_input("공동수급사 지분율(%)", value=0, min_value=0, max_value=100, step=5)

# ============================================================
# TAB 7. 하도급관리계획서 / 하도급통보
# ============================================================
with tabs[6]:
    st.subheader("🤝 하도급관리계획서 및 하도급통보")
    if "sub_df" not in st.session_state:
        st.session_state["sub_df"] = pd.DataFrame([
            {"하수급인 상호": "", "등록번호": "", "공종": "", "하도급금액": 0, "계약일": ""},
        ])
    sub_df = st.data_editor(st.session_state["sub_df"], num_rows="dynamic", key="sub_editor")
    st.session_state["sub_df"] = sub_df

    total_sub = sub_df["하도급금액"].sum()
    sub_ratio = total_sub / total_contract * 100 if total_contract else 0
    c1, c2 = st.columns(2)
    c1.metric("하도급 총액", fmt(total_sub))
    c2.metric("도급금액 대비 하도급 비율", f"{sub_ratio:.1f}%")

    if sub_ratio > 100:
        st.error("❌ 하도급 총액이 도급금액을 초과했습니다.")

    st.caption(
        "건설산업기본법 제29조: 하도급 계약 체결일로부터 30일 이내 발주자에게 통보(하도급통보서) 필요. "
        "일괄하도급(도급받은 공사 전부 또는 주요 부분 전부를 하도급) 금지 규정 위반 여부를 별도로 점검하세요."
    )

# ============================================================
# TAB 8. 착공계 패키지 다운로드 (원본 서식 그대로)
# ============================================================
with tabs[7]:
    st.subheader("📦 착공계 패키지 다운로드")
    st.write(
        "원본 착공계 서식(공문·표지·목차·착공신고서·현장대리인계·안전관리비·환경보전비·직접시공계획서 등)에 "
        "왼쪽 사이드바와 각 탭에서 입력한 값을 그대로 채워서 내려받습니다. 서식 레이아웃·수식은 원본과 동일합니다."
    )

    if not os.path.exists(TEMPLATE_PATH):
        st.error("템플릿 파일(smartsite_template_master.xlsx)을 찾을 수 없습니다. 앱 코드와 같은 폴더에 있어야 합니다.")
    elif st.button("🚀 착공계 패키지 생성하기", type="primary"):
        wb = load_workbook(TEMPLATE_PATH)

        # --- 마스터 시트 ---
        ws = wb["공사정보_출력X"]
        ws["C2"] = project_name
        ws["C3"] = work_category
        ws["C4"] = site_address
        ws["C5"] = datetime.combine(report_date, datetime.min.time())
        ws["C6"] = contractor_name
        ws["D6"] = contractor_tel
        ws["E6"] = contractor_addr
        ws["C7"] = ceo_name
        ws["C8"] = client_name
        ws["D8"] = client_tel
        ws["C9"] = total_contract
        ws["C10"] = contract_date_dt
        ws["C11"] = datetime.combine(start_date_master, datetime.min.time())
        ws["C12"] = datetime.combine(end_date_master, datetime.min.time())
        ws["C13"] = site_agent
        ws["E13"] = site_agent_tel
        ws["E14"] = site_agent_addr
        ws["E17"] = site_agent_license_no
        ws["E19"] = site_agent_grade
        ws["C21"] = contract_number

        # --- 산업안전보건관리비 ---
        ws_safety = wb["5-2.안전관리비사용계획서"]
        ws_safety["G11"] = user_safety_fee
        for i, v in enumerate(safety_usage):
            ws_safety.cell(row=14 + i, column=16, value=v)  # P열

        # --- 환경보전비 ---
        ws_env = wb["6-2.환경보전비사용계획서"]
        ws_env["G10"] = e_target
        for i, v in enumerate(env_usage):
            ws_env.cell(row=13 + i, column=15, value=v)  # O열

        # --- 직접시공내역서 (10-1) ---
        ws_dc = wb["10-1.직접시공내역서"]
        for i in range(6):
            row = 7 + i
            name = edited_df.iloc[i]["공종명"] if i < len(edited_df) else ""
            material = edited_df.iloc[i]["재료비"] if i < len(edited_df) else 0
            labor = edited_df.iloc[i]["노무비"] if i < len(edited_df) else 0
            expense = edited_df.iloc[i]["경비"] if i < len(edited_df) else 0
            ws_dc.cell(row=row, column=1, value=name or f"{i+1}. 공종 미입력")
            ws_dc.cell(row=row, column=5, value=int(material))
            ws_dc.cell(row=row, column=7, value=int(labor))
            ws_dc.cell(row=row, column=9, value=int(expense))

        # --- NUMBERSTRING 함수(LibreOffice 미지원) 결과를 파이썬 계산값으로 고정 ---
        words = amount_in_words(total_contract) if total_contract else ""
        for sheet_name, cell_ref in [
            ("1.착공신고서", "D4"),
            ("4.품질관리", "D4"),
            ("5. 안전관리", "D4"),
            ("6. 환경관리", "D4"),
        ]:
            if sheet_name in wb.sheetnames:
                wb[sheet_name][cell_ref] = words

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        st.success("✅ 착공계 패키지가 생성되었습니다. 아래 버튼으로 내려받으세요.")
        st.download_button(
            "📥 착공계 패키지 다운로드 (.xlsx)",
            data=buf,
            file_name=f"{project_name or '착공계'}_패키지.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption(
            "⚠️ 엑셀에서 열었을 때 수식이 즉시 재계산되지 않으면 Ctrl+Alt+F9(전체 재계산)를 눌러주세요."
        )

    st.divider()
    st.subheader("📦 기성청구서 패키지 다운로드")
    st.write(
        "원본 기성청구서 서식(기성공문·기성계·기성금청구서·기성부분검사원·선금정산서·표지 등)에 "
        "입력값을 채워서 내려받습니다. 공사원가계산서·집계표, 선금정산 상세내역(외주비·재료비·노무비·경비) "
        "등 세부 산출 시트는 프로젝트마다 달라 예시 1건만 넣어뒀습니다. 직접 채우셔야 합니다."
    )
    if not os.path.exists(GISUNG_TEMPLATE_PATH):
        st.error("템플릿 파일(gisung_template_master.xlsx)을 찾을 수 없습니다. 앱 코드와 같은 폴더에 있어야 합니다.")
    elif st.button("🚀 기성청구서 패키지 생성하기", type="primary"):
        wbg = load_workbook(GISUNG_TEMPLATE_PATH)
        wsg = wbg["기성정보_출력X"]

        wsg["C2"] = project_name
        wsg["C3"] = site_address
        wsg["C4"] = client_name
        wsg["C5"] = ""
        wsg["C6"] = contractor_name
        wsg["C7"] = contractor_addr
        wsg["C8"] = ceo_name
        wsg["C9"] = total_contract
        wsg["C10"] = contract_date_dt
        wsg["C11"] = datetime.combine(start_date_master, datetime.min.time())
        wsg["C12"] = datetime.combine(end_date_master, datetime.min.time())
        wsg["C13"] = None
        wsg["C14"] = current_progress_pct
        wsg["C15"] = prev_gisung_amount
        wsg["C16"] = progress_amount
        wsg["C17"] = advance_amount
        wsg["C18"] = labor_received
        wsg["C19"] = advance_deduction
        wsg["C20"] = gisung_deduction
        wsg["C21"] = bank_name
        wsg["C22"] = bank_account
        wsg["C23"] = is_jv
        wsg["C24"] = jv_name
        wsg["C25"] = jv_share

        contract_words = amount_in_words(total_contract) if total_contract else ""
        wbg["기성계"]["D4"] = contract_words
        wbg["기성금청구서"]["D4"] = contract_words
        wbg["선금정산서"]["A11"] = datetime.combine(advance_receipt_date, datetime.min.time())

        bufg = BytesIO()
        wbg.save(bufg)
        bufg.seek(0)

        st.success("✅ 기성청구서 패키지가 생성되었습니다. 아래 버튼으로 내려받으세요.")
        st.download_button(
            "📥 기성청구서 패키지 다운로드 (.xlsx)",
            data=bufg,
            file_name=f"{project_name or '기성청구'}_패키지.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption(
            "⚠️ 공사원가계산서·총괄집계표·공종별내역서 시트는 프로젝트별 산출내역이라 자동 채움 대상이 아닙니다. "
            "엑셀에서 열어 직접 입력해 주세요. 또한 수식이 즉시 재계산되지 않으면 Ctrl+Alt+F9를 눌러주세요."
        )

    st.divider()
    st.subheader("보조 서식 (하도급관리) 다운로드")
    st.caption("하도급관리계획서는 별도 원본 서식이 아직 없어 SmartSite 자체 표 형식으로 제공됩니다.")

    if st.button("📥 보조 서식 엑셀 생성하기"):
        wb2 = Workbook()
        wb2.remove(wb2.active)
        header_fill = PatternFill("solid", fgColor="1F5B54")
        header_font = Font(color="FFFFFF", bold=True)

        def write_sheet(title, rows):
            ws2 = wb2.create_sheet(title)
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            for c_idx in range(1, (len(rows[0]) if rows else 1) + 1):
                ws2.column_dimensions[get_column_letter(c_idx)].width = 22

        sub_rows = [list(sub_df.columns)] + sub_df.values.tolist()
        write_sheet("하도급관리", [["하도급총액", total_sub], ["하도급비율(%)", round(sub_ratio, 1)], []] + sub_rows)

        buf2 = BytesIO()
        wb2.save(buf2)
        buf2.seek(0)
        st.download_button(
            "📥 보조 서식 다운로드 (.xlsx)",
            data=buf2,
            file_name=f"{project_name or '현장'}_보조서식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
